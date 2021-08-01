import re
from enum import Enum, auto
from typing import Tuple, Union

from openpyxl.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

MAX_ROWS = 1048576
MAX_COLS = 16384

_POSITION_RE = re.compile(r'([^!]+)!\$([A-Z]+)\$(\d+)')
_COORD_RE = re.compile(r'([A-Z]+)(\d+)')
_RANGE_RE = re.compile(r'([A-Z\d]+):([A-Z\d]+)')


def _check_col_value(val: int):
    if not 1 <= val <= MAX_COLS:
        raise InvalidPositionError('Column index out of range: %d' % val)


def _check_row_value(val: int):
    if not 1 <= val <= MAX_ROWS:
        raise InvalidPositionError('Row index out of range: %d' % val)


def get_ws_min_coord(ws: Worksheet, axis: 'Axis') -> int:
    if axis == axis.COL:
        return ws.min_column
    return ws.min_row


def get_ws_max_coord(ws: Worksheet, axis: 'Axis') -> int:
    if axis == axis.COL:
        return ws.max_column
    return ws.max_row


class InvalidPositionError(Exception):
    pass


class InvalidRangeError(Exception):
    pass


class Axis(Enum):
    ROW = auto()
    COL = auto()


class Direction(Enum):
    UP = (0, -1)
    RIGHT = (1, 0)
    DOWN = (0, 1)
    LEFT = (-1, 0)

    @property
    def d_col(self) -> int:
        return self.value[0]

    @property
    def d_row(self) -> int:
        return self.value[1]

    @property
    def axis(self) -> Axis:
        if self.d_col != 0:
            return Axis.ROW
        return Axis.COL

    @property
    def opposite(self) -> 'Direction':
        return Direction(tuple([-x for x in self.value]))

    def __str__(self):
        return self.name.lower()


class Position:
    def __init__(self, *args):
        """
        Example: ``Position(2, 4)`` or ``Position('B4')``

        :param args: col: [int, str], row: [int, str] OR pos: str
        """
        if len(args) == 2:
            self.col = Position._parse_colval(args[0])
            self.row = Position._parse_rowval(args[1])
        elif len(args) == 1:
            m = _COORD_RE.match(args[0])

            if not m:
                raise InvalidPositionError('Invalid coord string: %s' %
                                           args[0])

            self.col = Position._parse_colval(m[1])
            self.row = Position._parse_rowval(m[2])
        else:
            raise TypeError('Position requires 1-2 positional arguments')

    @staticmethod
    def _parse_colval(val: Union[int, str]) -> int:
        if not isinstance(val, int):
            try:
                val = column_index_from_string(val)
            except ValueError:
                raise InvalidPositionError('Invalid column index: %s' %
                                           str(val))

        _check_col_value(val)
        return val

    @staticmethod
    def _parse_rowval(val: Union[int, str]) -> int:
        if not isinstance(val, int):
            try:
                val = int(val)
            except ValueError:
                raise InvalidPositionError('Invalid row index: %s' % str(val))

        _check_row_value(val)
        return val

    @classmethod
    def from_abs_string(cls, abs_str: str) -> Tuple['Position', str]:
        """
        Get position and sheet name from absolute position string.

        Example: ``Sheet1!$B$4``

        :param abs_str: Absolute position string
        :return: (Position, Sheet name)
        """
        m = _POSITION_RE.match(abs_str)

        if not m:
            raise InvalidPositionError('Invalid position string: %s' % abs_str)

        return cls(m[2], m[3]), m[1]

    def shifted(self, direction: Direction, d: int = 1) -> 'Position':
        """
        Return a new position object that is shifted by ``d``
        cells in the given direction.

        :param direction: Shift direction
        :param d: Distance in cells
        :return: New position
        """
        ncol = self.col + direction.value[0] * d
        nrow = self.row + direction.value[1] * d

        return Position(ncol, nrow)

    def dir_distance(self, pos_b: 'Position', direction: Direction):
        """
        Calculate the distance between to positions in a given
        direction.

        ::

          pos_a.shifted(dir, pos_a.dir_distance(pos_b, dir)) == pos_b

        :param pos_b: Second position
        :param direction: Direction
        :return: Distance in cells
        """
        if direction.axis == Axis.COL:
            return (pos_b.row - self.row) * direction.d_row
        return (pos_b.col - self.col) * direction.d_col

    def combine(self, axis: Axis, pos_b: Union['Position', int]):
        """
        Return a new position object with the ``axis`` coordinate
        of this position and the other coordinate of pos_b.

        :param axis: Axis (Row/Col)
        :param pos_b: Second position OR other coordinate number
        :return: New position
        """
        if isinstance(pos_b, int):
            pos_b_data = (pos_b, pos_b)
        else:
            pos_b_data = (pos_b.row, pos_b.col)

        if axis == Axis.ROW:
            return Position(pos_b_data[1], self.row)
        return Position(self.col, pos_b_data[0])

    def is_in(self, ws: Worksheet) -> bool:
        """
        Check if the position is within the data containing area of
        a worksheet.

        :param ws: OpenPyXL Worksheet
        :return: is_in
        """
        return ws.min_row <= self.row <= ws.max_row and \
            ws.min_column <= self.col <= ws.max_column

    def get_cell(self, ws: Worksheet) -> Cell:
        """
        Get the cell of a worksheet located at the position

        :param ws: OpenPyXL Worksheet
        :return: OpenPyXL Cell
        """
        return ws.cell(self.row, self.col)

    def is_cell_empty(self, ws: Worksheet) -> bool:
        """
        Check if the cell at this position in the given worksheet
        is empty.

        :param ws: OpenPyXL Worksheet
        :return: is_empty
        """
        cell = self.get_cell(ws)
        return cell.value is None and cell.comment is None and \
            not cell.has_style

    def get_coord(self, axis: Axis) -> int:
        """
        Return the position's coordinate of the given axis

        :param axis: Axis (ROW/COL)
        :return: Coordinate
        """
        if axis == Axis.COL:
            return self.col
        return self.row

    def __eq__(self, other):
        return self.row == other.row and self.col == other.col

    def __str__(self):
        return get_column_letter(self.col) + str(self.row)

    def __repr__(self):
        return '<Position: %s>' % str(self)


class Range:
    def __init__(self, min_row: int, max_row: int, min_col: int, max_col: int):
        self.min_row = min_row
        self.max_row = max_row
        self.min_col = min_col
        self.max_col = max_col

        self._verify()

    @classmethod
    def from_pos(cls, pos_a: Position, pos_b: Position) -> 'Range':
        return cls(min(pos_a.row, pos_b.row), max(pos_a.row, pos_b.row),
                   min(pos_a.col, pos_b.col), max(pos_a.col, pos_b.col))

    @classmethod
    def from_str(cls, range_str: str):
        m = _RANGE_RE.match(range_str)

        if not m:
            raise InvalidRangeError('Invalid range string: ' + range_str)

        return cls.from_pos(Position(m[1]), Position(m[2]))

    def _verify(self):
        # Check if row/column values are within allowed range
        _check_row_value(self.min_row)
        _check_row_value(self.max_row)
        _check_col_value(self.min_col)
        _check_col_value(self.max_col)

        # Assert that min values are smaller than max values (or equal)
        if self.min_row > self.max_row:
            raise InvalidRangeError('min_row larger than max_row')
        if self.min_col > self.max_col:
            raise InvalidRangeError('min_col larger than max_col')

    def is_inside(self, pos: Position) -> bool:
        """
        Check if a position is located inside the range.

        :param pos: Position to check
        :return: True if position is inside range
        """
        return self.min_row <= pos.row <= self.max_row and \
            self.min_col <= pos.col <= self.max_col

    def extended(self, direction: Direction, n: int = 1) -> 'Range':
        """
        Return a new range object extended in the specified direction

        :param direction: Direction to extend to
        :param n: Amount of cells
        :raise InvalidRangeError: if the range is extended outside the valid
        space
        :return: New range object
        """
        # Adding to minimum / subtracting from maximum values
        # will result in invalid data, so invert distance and direction
        # in case of negative distance values
        if n < 0:
            n = -n
            direction = direction.opposite

        cp = Range(self.min_row, self.max_row, self.min_col, self.max_col)

        if direction == Direction.UP:
            cp.min_row -= n
        elif direction == Direction.DOWN:
            cp.max_row += n
        elif direction == Direction.LEFT:
            cp.min_col -= n
        elif direction == Direction.RIGHT:
            cp.max_col += n

        cp._verify()
        return cp

    def __eq__(self, other):
        return self.min_row == other.min_row and \
               self.max_row == other.max_row and \
               self.min_col == other.min_col and \
               self.max_col == other.max_col

    def __str__(self):
        return '%s%d:%s%d' % (get_column_letter(self.min_col), self.min_row,
                              get_column_letter(self.max_col), self.max_row)

    def __repr__(self):
        return '<Range: %s>' % str(self)


class MarkerAbs:
    def get_position(self, ws: Worksheet) -> Position:
        pass

    def get_cell(self, ws: Worksheet) -> Cell:
        return self.get_position(ws).get_cell(ws)


class MarkerPos(MarkerAbs):
    def __init__(self, *args):
        self.pos = Position(*args)

    def get_position(self, ws: Worksheet) -> Position:
        return self.pos


class MarkerName(MarkerAbs):
    def __init__(self, name: str):
        self.name = name

    def get_position(self, ws: Worksheet) -> Position:
        dn = ws.parent.defined_names.get(self.name)
        if dn:
            # Parse defined name
            pos, sheet_name = Position.from_abs_string(dn.value)

            if sheet_name != ws.title:
                raise InvalidPositionError('Marker %s not in worksheet %s' %
                                           (self.name, sheet_name))

            return pos
        else:
            raise InvalidPositionError('Defined name %s not found' % self.name)


class MarkerPattern(MarkerAbs):
    def __init__(self, initial_marker: MarkerAbs, pattern: Union[str,
                                                                 re.Pattern],
                 direction: Direction, max_range: int):
        self.initial_marker = initial_marker
        self.direction = direction
        self.max_range = max_range
        self.rex = re.compile(pattern)

    def get_position(self, ws: Worksheet) -> Position:
        initial_pos = self.initial_marker.get_position(ws)

        for d in range(self.max_range + 1):
            pos = initial_pos.shifted(self.direction, d)
            val = str(pos.get_cell(ws).value)

            if self.rex.search(val):
                return pos

        raise InvalidPositionError(
            'Cell matching pattern %s max. %d cells %s from %s not found' %
            (str(self.rex), self.max_range, str(
                self.direction), str(initial_pos)))
