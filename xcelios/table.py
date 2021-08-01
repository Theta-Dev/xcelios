import re
from datetime import datetime
from typing import Any, Dict, Type

from openpyxl.worksheet.worksheet import Worksheet

from xcelios.position import Axis, Direction, MarkerAbs, Position, Range


def insert_rows_cols_withref(ws: Worksheet,
                             index: int,
                             axis: Axis,
                             n: int = 1):
    """
    Insert the specified amount of rows or columns after the given index.

    TODO: The ``move_range`` OpenPyXL function does not update references
          outside the moved range. For these another solution is needed.

    :param ws: OpenPyXL worksheet
    :param index: Row/column index
    :param axis: Axis (ROW/COL)
    :param n: Amount of rows/columns
    """
    if axis == Axis.ROW:
        rg = Range(index, ws.max_row, ws.min_column, ws.max_column)
        ws.move_range(str(rg), rows=n, translate=True)
    else:
        rg = Range(ws.min_row, ws.max_row, index, ws.max_column)
        ws.move_range(str(rg), cols=n, translate=True)


def delete_rows_cols_withref(ws: Worksheet,
                             index: int,
                             axis: Axis,
                             n: int = 1):
    """
    Delete the specified amount of rows or columns after the given index.

    :param ws: OpenPyXL worksheet
    :param index: Row/column index
    :param axis: Axis (ROW/COL)
    :param n: Amount of rows/columns
    """
    insert_rows_cols_withref(ws, index + 1, axis, -n)


class TableParseError(Exception):
    pass


class Table:
    def __init__(self,
                 ws: Worksheet,
                 initial_marker: MarkerAbs,
                 obj_class: Type,
                 header_dir: Direction = Direction.RIGHT,
                 body_dir: Direction = Direction.DOWN,
                 max_blanks: int = 1):
        self.ws = ws
        self.initial_pos = initial_marker.get_position(self.ws)
        self.obj_class = obj_class
        self.header_dir = header_dir
        self.body_dir = body_dir

        # Maximum number of blank rows/cols to ignore
        self.max_blanks = max_blanks

        self.title_positions: Dict[str, Position] = dict()
        self.title_range = Range.from_pos(self.initial_pos, self.initial_pos)
        self.datasets = []
        self.final_pos = self.initial_pos

        self._locate_headers()

    def _locate_headers(self):
        # Read type annotations from dataclass
        title_rexes = dict()
        for key in self.obj_class.__annotations__.keys():
            title_rexes[key] = re.compile(
                re.escape(key).replace('_', r'[_\- ]?'), re.IGNORECASE)

        blanks = 0
        pos = self.initial_pos
        last_valid_pos = self.initial_pos

        # Stop iteration after encountering more than max_blanks empty cells
        # after eachother, reaching the end of the worksheet
        # or having found all titles
        while blanks <= self.max_blanks and pos.is_in(self.ws) and title_rexes:
            val = pos.get_cell(self.ws).value

            if val:
                blanks = 0
                found_key = None

                for key, rex in title_rexes.items():
                    if rex.match(str(val)):
                        found_key = key
                        break

                if found_key is not None:
                    self.title_positions[found_key] = pos
                    title_rexes.pop(found_key)
                    last_valid_pos = pos
            else:
                blanks += 1

            pos = pos.shifted(self.header_dir)

        if title_rexes:
            raise TableParseError('Could not find table headers: %s' %
                                  ', '.join(title_rexes.keys()))

        # Get title range
        self.title_range = Range.from_pos(self.initial_pos, last_valid_pos)

    @staticmethod
    def _cast(val: Any, typ: Type) -> Any:
        # Special case: None
        if val is None:
            if typ is str:
                return ''
            if typ is int:
                return 0

        # Special case: datetime
        if typ == datetime:
            if isinstance(val, datetime):
                return val
            else:
                # TODO: Parse date strings
                return None

        try:
            return typ(val)
        except TypeError:
            return None

    def read_datasets(self):
        self.datasets = []

        line = 1
        blanks = 0

        while blanks <= self.max_blanks and self.initial_pos.shifted(
                self.body_dir, line).is_in(self.ws):
            is_blank = True
            data = dict()

            # Fetch data for the new dataset
            for key, tpos in self.title_positions.items():
                pos = tpos.shifted(self.body_dir, line)
                raw_val = pos.get_cell(self.ws).value

                if raw_val is not None:
                    is_blank = False

                val = Table._cast(raw_val, self.obj_class.__annotations__[key])
                data[key] = val

            if is_blank:
                blanks += 1
            else:
                # Create new dataset
                self.datasets.append(self.obj_class(**data))
                self.final_pos = self.initial_pos.shifted(self.body_dir, line)

            line += 1

    @property
    def initial_length(self) -> int:
        """Return the initial length"""
        return self.initial_pos.dir_distance(self.final_pos, self.body_dir)

    @property
    def table_range(self) -> Range:
        return self.title_range.extended(self.body_dir, self.initial_length)

    def _get_space(self) -> int:
        """
        Determine amount of whitespace between end of table and the
        following content

        :return: Number of empty rows/cols
        """
        space = None

        for tpos in self.title_positions.values():
            pos = tpos.combine(self.body_dir.axis,
                               self.final_pos).shifted(self.body_dir)
            s = 0
            while pos.is_in(self.ws) and pos.is_cell_empty(self.ws):
                pos = pos.shifted(self.body_dir)
                s += 1

            if not pos.is_cell_empty(self.ws):
                if space is None:
                    space = s
                else:
                    space = min(space, s)

        return space

    def _is_row_empty(self, n_row: int) -> bool:
        """
        Check if a row does not contain any cells with content
        (except for the table's own cells)

        :param n_row: Row number
        :return: True if empty
        """
        t_range = self.table_range

        if self.body_dir.axis == Axis.COL:
            pos = Position(self.ws.min_column, n_row)
        else:
            pos = Position(n_row, self.ws.min_row)

        while pos.is_in(self.ws):
            if not t_range.is_inside(pos) and not pos.is_cell_empty(self.ws):
                return False

            pos = pos.shifted(self.header_dir)

        return True

    def _adjust_space(self, new_n_rows: int):
        """
        move other cells in the worksheets so that the space for the
        table matches the number of datasets.

        updates ``final_pos`` and ``original_length``.
        """
        diff_rows = new_n_rows - self.initial_length

        # If no space adjustments have to be made, exit
        if diff_rows == 0:
            return

        n_rows = abs(diff_rows)
        i_pos = self.final_pos

        while True:
            # If we didn't find non-empty cells after the table, exit
            if not i_pos.is_in(self.ws):
                return

            next_pos = i_pos.shifted(self.body_dir)

            # If the next cell below the table contains data, exit the loop
            # This our starting position for the extending/shrinking
            if not next_pos.is_cell_empty(self.ws):
                break
            i_pos = next_pos

        if diff_rows > 0:
            self._insert_space(i_pos, n_rows)
        else:
            self._remove_space(i_pos, n_rows)

    def _insert_space(self, i_pos: Position, n_rows: int):
        # Find a completely empty row where the new rows will be
        # inserted
        pos = i_pos

        while pos != self.initial_pos:
            row = pos.get_coord(self.header_dir.axis)
            if self._is_row_empty(row):
                insert_rows_cols_withref(self.ws, row, self.header_dir.axis,
                                         n_rows)
                return

            pos = pos.shifted(self.body_dir.opposite)

        raise Exception('Could not insert %d rows' % n_rows)

    def _remove_space(self, i_pos: Position, n_rows: int):
        # Look for completely empty rows do delete
        # Preserve the whitespace below the table
        # Stop when encountering a non-empty row to prevent
        # destroying the table layout
        pos = i_pos.shifted(self.body_dir.opposite, self._get_space())

        while pos != self.initial_pos and n_rows > 0:
            row = pos.get_coord(self.header_dir.axis)
            if self._is_row_empty(row):
                delete_rows_cols_withref(self.ws, row, self.header_dir.axis)
                n_rows -= 1
            else:
                break

            pos = pos.shifted(self.body_dir.opposite)

        if n_rows != 0:
            raise Exception('Could not insert %d rows' % n_rows)

    def write_datasets(self):
        self._adjust_space(len(self.datasets))

        r_pos = self.initial_pos

        for d in self.datasets:
            r_pos = r_pos.shifted(self.body_dir)

            for key, hpos in self.title_positions.items():
                pos = r_pos.combine(self.header_dir.axis, hpos)
                pos.get_cell(self.ws).value = getattr(d, key, None)
